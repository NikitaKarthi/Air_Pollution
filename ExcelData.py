import openpyxl
from openpyxl import load_workbook
import pandas as pd

workbook = load_workbook('Kanpur.xlsx')
listSheets = workbook.sheetnames
sheet1 = workbook[listSheets[0]]

dfJan19 = pd.DataFrame(sheet1.values)
tempDF = dfJan19.loc[13:16, [0,1,3]]

print(tempDF)
'''
dfJan19PM25 = pd.DataFrame()

dfJan19PM25 = dfJan19PM25.append(tempDF)

i = 17
while (dfJan19.loc[i,0] != None):
    tempDF = dfJan21.loc[i, 0:2]
    dfJan19PM25 = dfJan19PM25.append(tempDF)
    i = i + 1

dfJan19PM25.to_csv('Jan19PM25')
print("HEllo")


'''  
#print(dfJan19.loc[16,0] == 'From Date')
#for row in sheet1.iter_rows(min_row=1, max_col=3, max_row=5, values_only=True):
#    print(row)
